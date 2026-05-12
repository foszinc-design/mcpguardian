from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook

from guardian.atomic_io import load_json
from guardian.validators.xlsx_validator import generate_xlsx_artifacts


def make_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Month", "Revenue", "Cost"])
    ws.append(["Jan", 100, 40])
    ws.append(["Feb", 130, 50])
    ws["D2"] = "=B2-C2"
    ws.merge_cells("E1:F1")
    ws.auto_filter.ref = "A1:D3"

    hidden = wb.create_sheet("HiddenRaw")
    hidden.sheet_state = "hidden"
    hidden.append(["id", "value"])
    hidden.append([1, 10])
    hidden.append([2, 20])

    wb.save(path)


class XlsxValidatorTests(unittest.TestCase):
    def test_generates_sheet_inventory_and_row_summary(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "sample.xlsx"
            make_workbook(workbook)
            outputs = generate_xlsx_artifacts(workbook, tmp)

            inventory = load_json(outputs["sheet_inventory"])
            row_summary = load_json(outputs["row_count_summary"])

            self.assertEqual(inventory["artifact_type"], "sheet_inventory")
            self.assertEqual(inventory["sheet_count"], 2)
            self.assertEqual(inventory["hidden_sheet_count"], 1)
            self.assertEqual(inventory["hidden_sheets"], ["HiddenRaw"])
            self.assertEqual(row_summary["artifact_type"], "row_count_summary")
            self.assertGreaterEqual(row_summary["workbook_total_non_empty_rows"], 5)

    def test_default_coverage_is_not_safe_for_global_claims(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "sample.xlsx"
            make_workbook(workbook)
            outputs = generate_xlsx_artifacts(workbook, tmp)
            coverage = load_json(outputs["coverage_report"])

            self.assertEqual(coverage["artifact_type"], "coverage_report")
            self.assertFalse(coverage["global_claim_safe"])
            self.assertEqual(coverage["coverage_basis"], "metadata_inventory_only")
            self.assertIn("HiddenRaw", coverage["missing_sheets"])

    def test_assume_full_analysis_marks_global_claim_safe(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook = Path(tmp) / "sample.xlsx"
            make_workbook(workbook)
            outputs = generate_xlsx_artifacts(workbook, tmp, assume_full_analysis=True)
            coverage = load_json(outputs["coverage_report"])
            result = load_json(outputs["validator_result"])

            self.assertTrue(coverage["global_claim_safe"])
            self.assertEqual(coverage["sheet_coverage_ratio"], 1.0)
            self.assertTrue(result["ok"])
            self.assertIn("Workbook contains hidden or veryHidden sheets.", result["warnings"])


if __name__ == "__main__":
    unittest.main()
