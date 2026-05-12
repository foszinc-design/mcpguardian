"""Document-specialized native tools for MCPGuardian Phase 8B.

These tools intentionally provide conservative metadata/excerpt extraction. They
are not a replacement for full Office/PDF semantic parsing; they give the gate
and operator enough structured evidence to avoid blind binary-file handling.
"""
from __future__ import annotations

import re
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from ..atomic_io import sha256_file
from .common import ToolContext, emit, error_code_for_exception, resolve_path, tool_error, tool_ok


def inspect_xlsx(ctx: ToolContext, arguments: dict[str, Any]) -> dict[str, Any]:
    try:
        path = resolve_path(ctx, str(arguments.get("path") or ""), must_exist=True)
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            return tool_error(ctx, "UNSUPPORTED_FILE_TYPE", "guardian_inspect_xlsx supports .xlsx/.xlsm only")
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            return tool_error(ctx, "MISSING_DEPENDENCY", f"openpyxl is required: {exc}")
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = []
        total_rows = 0
        for ws in wb.worksheets:
            non_empty_rows = 0
            sample_rows = []
            max_sample_rows = int(arguments.get("sample_rows") or 5)
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                values = [cell for cell in row]
                if any(cell is not None for cell in values):
                    non_empty_rows += 1
                    if len(sample_rows) < max_sample_rows:
                        sample_rows.append([_safe_cell(cell) for cell in values[:20]])
                if idx > 100000 and not bool(arguments.get("scan_all", False)):
                    break
            total_rows += non_empty_rows
            sheets.append({
                "name": ws.title,
                "visible": ws.sheet_state == "visible",
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "non_empty_rows_sampled": non_empty_rows,
                "sample_rows": sample_rows,
            })
        wb.close()
        data = {"path": str(path), "sha256": sha256_file(path), "sheet_count": len(sheets), "sheets": sheets, "total_non_empty_rows_sampled": total_rows}
        emit(ctx, "native_document_xlsx_inspected", path=str(path), sheet_count=len(sheets))
        return tool_ok(ctx, message="xlsx inspected", data=data)
    except Exception as exc:
        return tool_error(ctx, error_code_for_exception(exc), str(exc))


def read_docx(ctx: ToolContext, arguments: dict[str, Any]) -> dict[str, Any]:
    try:
        path = resolve_path(ctx, str(arguments.get("path") or ""), must_exist=True)
        if path.suffix.lower() != ".docx":
            return tool_error(ctx, "UNSUPPORTED_FILE_TYPE", "guardian_read_docx supports .docx only")
        max_chars = int(arguments.get("max_chars") or 20000)
        text = _extract_docx_text(path)
        truncated = len(text) > max_chars
        excerpt = text[:max_chars]
        data = {"path": str(path), "sha256": sha256_file(path), "text": excerpt, "total_chars": len(text), "truncated": truncated}
        emit(ctx, "native_document_docx_read", path=str(path), chars=len(excerpt))
        return tool_ok(ctx, message="docx read", data=data)
    except Exception as exc:
        return tool_error(ctx, error_code_for_exception(exc), str(exc))


def inspect_pdf(ctx: ToolContext, arguments: dict[str, Any]) -> dict[str, Any]:
    try:
        path = resolve_path(ctx, str(arguments.get("path") or ""), must_exist=True)
        if path.suffix.lower() != ".pdf":
            return tool_error(ctx, "UNSUPPORTED_FILE_TYPE", "guardian_inspect_pdf supports .pdf only")
        raw = path.read_bytes()
        page_count = len(re.findall(rb"/Type\s*/Page\b", raw))
        metadata = _extract_pdf_info(raw)
        # Conservative text hint extraction. Full PDF text extraction is out of scope
        # unless a dedicated parser is introduced later.
        text_hints = [m.decode("latin1", errors="ignore") for m in re.findall(rb"\(([^()]{1,200})\)\s*Tj", raw)[:50]]
        data = {
            "path": str(path),
            "sha256": sha256_file(path),
            "size_bytes": path.stat().st_size,
            "page_count_estimate": page_count,
            "metadata": metadata,
            "text_hints": text_hints,
            "full_text_extracted": False,
        }
        emit(ctx, "native_document_pdf_inspected", path=str(path), page_count_estimate=page_count)
        return tool_ok(ctx, message="pdf inspected", data=data)
    except Exception as exc:
        return tool_error(ctx, error_code_for_exception(exc), str(exc))


def _safe_cell(value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _extract_docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs = []
    for para in root.findall(".//w:p", ns):
        pieces = [node.text or "" for node in para.findall(".//w:t", ns)]
        if pieces:
            paragraphs.append("".join(pieces))
    return "\n".join(paragraphs)


def _extract_pdf_info(raw: bytes) -> dict[str, str]:
    out: dict[str, str] = {}
    for key in ["Title", "Author", "Subject", "Creator", "Producer"]:
        pattern = rb"/" + key.encode("ascii") + rb"\s*\(([^)]{0,500})\)"
        match = re.search(pattern, raw)
        if match:
            out[key.lower()] = match.group(1).decode("latin1", errors="replace")
    return out
