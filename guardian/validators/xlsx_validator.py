"""XLSX metadata artifact generator for MCPGuardian Phase 2.

This validator intentionally does not infer business meaning from workbook data.
It creates structural evidence that downstream gates and report validators can
use to reject partial-workbook conclusions.

Artifacts produced:
- sheet_inventory.json
- row_count_summary.json
- coverage_report.json
- validator_result.json
"""
from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..atomic_io import locked_atomic_write_json, sha256_file
from ..schemas import utc_now_iso

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}


@dataclass(frozen=True)
class SheetStats:
    name: str
    state: str
    visible: bool
    max_row: int
    max_column: int
    dimension: str
    non_empty_rows: int
    non_empty_cells: int
    first_non_empty_row: int | None
    last_non_empty_row: int | None
    first_non_empty_column: int | None
    last_non_empty_column: int | None
    merged_cells: int
    tables: int
    filters: bool
    auto_filter_ref: str | None
    formulas: int
    comments: int
    hidden_rows: int
    hidden_columns: int
    freeze_panes: str | None

    def to_inventory_dict(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "state": self.state,
            "visible": self.visible,
            "max_row": self.max_row,
            "max_column": self.max_column,
            "dimension": self.dimension,
            "non_empty_rows": self.non_empty_rows,
            "non_empty_cells": self.non_empty_cells,
            "first_non_empty_row": self.first_non_empty_row,
            "last_non_empty_row": self.last_non_empty_row,
            "first_non_empty_column": self.first_non_empty_column,
            "last_non_empty_column": self.last_non_empty_column,
            "merged_cells": self.merged_cells,
            "tables": self.tables,
            "filters": self.filters,
            "auto_filter_ref": self.auto_filter_ref,
            "formulas": self.formulas,
            "comments": self.comments,
            "hidden_rows": self.hidden_rows,
            "hidden_columns": self.hidden_columns,
            "freeze_panes": self.freeze_panes,
        }

    def to_row_summary_dict(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "state": self.state,
            "visible": self.visible,
            "max_row": self.max_row,
            "non_empty_rows": self.non_empty_rows,
            "non_empty_cells": self.non_empty_cells,
            "first_non_empty_row": self.first_non_empty_row,
            "last_non_empty_row": self.last_non_empty_row,
        }


def _cell_has_value(value: Any) -> bool:
    return value is not None and value != ""


def _sheet_stats(ws: Any) -> SheetStats:
    non_empty_rows = 0
    non_empty_cells = 0
    first_row: int | None = None
    last_row: int | None = None
    first_col: int | None = None
    last_col: int | None = None
    formulas = 0
    comments = 0

    for row in ws.iter_rows():
        row_has_value = False
        for cell in row:
            value = cell.value
            if _cell_has_value(value):
                row_has_value = True
                non_empty_cells += 1
                first_row = cell.row if first_row is None else min(first_row, cell.row)
                last_row = cell.row if last_row is None else max(last_row, cell.row)
                first_col = cell.column if first_col is None else min(first_col, cell.column)
                last_col = cell.column if last_col is None else max(last_col, cell.column)
                if isinstance(value, str) and value.startswith("="):
                    formulas += 1
            if cell.comment is not None:
                comments += 1
        if row_has_value:
            non_empty_rows += 1

    hidden_rows = sum(1 for dim in ws.row_dimensions.values() if getattr(dim, "hidden", False))
    hidden_columns = sum(1 for dim in ws.column_dimensions.values() if getattr(dim, "hidden", False))
    auto_filter_ref = str(ws.auto_filter.ref) if ws.auto_filter and ws.auto_filter.ref else None
    freeze_panes = None
    if ws.freeze_panes is not None:
        freeze_panes = str(ws.freeze_panes.coordinate if hasattr(ws.freeze_panes, "coordinate") else ws.freeze_panes)

    return SheetStats(
        name=ws.title,
        state=ws.sheet_state,
        visible=ws.sheet_state == "visible",
        max_row=ws.max_row,
        max_column=ws.max_column,
        dimension=ws.calculate_dimension(),
        non_empty_rows=non_empty_rows,
        non_empty_cells=non_empty_cells,
        first_non_empty_row=first_row,
        last_non_empty_row=last_row,
        first_non_empty_column=first_col,
        last_non_empty_column=last_col,
        merged_cells=len(ws.merged_cells.ranges),
        tables=len(ws.tables),
        filters=auto_filter_ref is not None,
        auto_filter_ref=auto_filter_ref,
        formulas=formulas,
        comments=comments,
        hidden_rows=hidden_rows,
        hidden_columns=hidden_columns,
        freeze_panes=freeze_panes,
    )


def _workbook_metadata(path: Path) -> dict[str, Any]:
    stat = path.stat()
    return {
        "path": str(path),
        "file_name": path.name,
        "extension": path.suffix.lower(),
        "size_bytes": stat.st_size,
        "sha256": sha256_file(path),
        "modified_at_epoch": stat.st_mtime,
    }


def inspect_workbook(path: str | Path) -> tuple[dict[str, Any], list[SheetStats]]:
    workbook_path = Path(path)
    if workbook_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported workbook extension: {workbook_path.suffix}. Supported: {sorted(SUPPORTED_EXTENSIONS)}")
    if not workbook_path.exists():
        raise FileNotFoundError(str(workbook_path))

    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    wb = load_workbook(workbook_path, read_only=False, data_only=False, keep_vba=keep_vba)
    try:
        stats = [_sheet_stats(ws) for ws in wb.worksheets]
        metadata = _workbook_metadata(workbook_path)
        metadata.update(
            {
                "workbook_title": wb.properties.title,
                "sheet_count": len(wb.worksheets),
                "defined_names_count": len(list(wb.defined_names.values())) if hasattr(wb.defined_names, "values") else 0,
                "created_at": wb.properties.created.isoformat() if wb.properties.created else None,
                "modified_at": wb.properties.modified.isoformat() if wb.properties.modified else None,
            }
        )
        return metadata, stats
    finally:
        wb.close()


def build_sheet_inventory(workbook: dict[str, Any], sheets: list[SheetStats]) -> dict[str, Any]:
    hidden = [sheet.name for sheet in sheets if sheet.state != "visible"]
    return {
        "artifact_type": "sheet_inventory",
        "generated_at": utc_now_iso(),
        "workbook": workbook,
        "sheet_count": len(sheets),
        "visible_sheet_count": sum(1 for sheet in sheets if sheet.visible),
        "hidden_sheet_count": len(hidden),
        "hidden_sheets": hidden,
        "sheets": [sheet.to_inventory_dict() for sheet in sheets],
    }


def build_row_count_summary(workbook: dict[str, Any], sheets: list[SheetStats]) -> dict[str, Any]:
    total_non_empty_rows = sum(sheet.non_empty_rows for sheet in sheets)
    total_non_empty_cells = sum(sheet.non_empty_cells for sheet in sheets)
    return {
        "artifact_type": "row_count_summary",
        "generated_at": utc_now_iso(),
        "workbook": {
            "file_name": workbook["file_name"],
            "sha256": workbook["sha256"],
        },
        "sheet_count": len(sheets),
        "workbook_total_non_empty_rows": total_non_empty_rows,
        "workbook_total_non_empty_cells": total_non_empty_cells,
        "visible_total_non_empty_rows": sum(sheet.non_empty_rows for sheet in sheets if sheet.visible),
        "hidden_total_non_empty_rows": sum(sheet.non_empty_rows for sheet in sheets if not sheet.visible),
        "sheets": [sheet.to_row_summary_dict() for sheet in sheets],
    }


def _normalize_sheet_selection(raw_names: list[str], all_sheet_names: list[str]) -> set[str]:
    if not raw_names:
        return set()
    selected: set[str] = set()
    for raw in raw_names:
        if raw in {"*", "ALL", "all"}:
            selected.update(all_sheet_names)
        elif raw in all_sheet_names:
            selected.add(raw)
        else:
            raise ValueError(f"Unknown sheet in --analyzed-sheet: {raw}")
    return selected


def build_coverage_report(
    workbook: dict[str, Any],
    sheets: list[SheetStats],
    *,
    analyzed_sheets: list[str] | None = None,
    assume_full_analysis: bool = False,
) -> dict[str, Any]:
    all_names = [sheet.name for sheet in sheets]
    total_rows = sum(sheet.non_empty_rows for sheet in sheets)
    total_cells = sum(sheet.non_empty_cells for sheet in sheets)

    if assume_full_analysis:
        selected = set(all_names)
        coverage_basis = "explicit_assume_full_analysis"
    else:
        selected = _normalize_sheet_selection(analyzed_sheets or [], all_names)
        coverage_basis = "declared_analyzed_sheets" if selected else "metadata_inventory_only"

    selected_sheets = [sheet for sheet in sheets if sheet.name in selected]
    covered_rows = sum(sheet.non_empty_rows for sheet in selected_sheets)
    covered_cells = sum(sheet.non_empty_cells for sheet in selected_sheets)
    missing_sheets = [sheet.name for sheet in sheets if sheet.name not in selected]
    hidden_missing_sheets = [sheet.name for sheet in sheets if not sheet.visible and sheet.name not in selected]

    sheet_coverage_ratio = 1.0 if not sheets else len(selected_sheets) / len(sheets)
    row_coverage_ratio = 1.0 if total_rows == 0 else covered_rows / total_rows
    cell_coverage_ratio = 1.0 if total_cells == 0 else covered_cells / total_cells
    global_claim_safe = bool(sheets) and not missing_sheets and covered_rows == total_rows and covered_cells == total_cells

    limitations: list[str] = []
    if coverage_basis == "metadata_inventory_only":
        limitations.append("No analyzed sheets were declared. This artifact proves workbook inventory coverage, not analytical coverage.")
    if hidden_missing_sheets:
        limitations.append("One or more hidden sheets were not declared as analyzed.")
    if missing_sheets:
        limitations.append("Workbook-level/global claims are unsafe until every sheet is included or explicitly excluded with rationale.")
    if assume_full_analysis:
        limitations.append("Full analysis coverage was asserted by CLI flag; downstream claim validation should still verify calculations.")

    return {
        "artifact_type": "coverage_report",
        "generated_at": utc_now_iso(),
        "workbook": {
            "file_name": workbook["file_name"],
            "sha256": workbook["sha256"],
        },
        "coverage_basis": coverage_basis,
        "global_claim_safe": global_claim_safe,
        "sheet_coverage_ratio": sheet_coverage_ratio,
        "row_coverage_ratio": row_coverage_ratio,
        "cell_coverage_ratio": cell_coverage_ratio,
        "total_sheets": len(sheets),
        "covered_sheets": len(selected_sheets),
        "missing_sheets": missing_sheets,
        "hidden_missing_sheets": hidden_missing_sheets,
        "total_non_empty_rows": total_rows,
        "covered_non_empty_rows": covered_rows,
        "total_non_empty_cells": total_cells,
        "covered_non_empty_cells": covered_cells,
        "analyzed_sheets": sorted(selected),
        "limitations": limitations,
    }


def build_validator_result(
    sheet_inventory: dict[str, Any],
    row_summary: dict[str, Any],
    coverage_report: dict[str, Any],
) -> dict[str, Any]:
    warnings: list[str] = []
    errors: list[str] = []

    if sheet_inventory["sheet_count"] == 0:
        errors.append("Workbook has no worksheets.")
    if sheet_inventory["hidden_sheet_count"] > 0:
        warnings.append("Workbook contains hidden or veryHidden sheets.")
    if not coverage_report["global_claim_safe"]:
        warnings.append("Coverage report does not support workbook-level/global claims.")
    if row_summary["workbook_total_non_empty_rows"] == 0:
        warnings.append("Workbook appears to contain no non-empty rows.")

    return {
        "artifact_type": "validator_result",
        "generated_at": utc_now_iso(),
        "validator": "xlsx_validator",
        "ok": not errors,
        "errors": errors,
        "warnings": warnings,
        "produced_artifacts": [
            "sheet_inventory.json",
            "row_count_summary.json",
            "coverage_report.json",
        ],
    }


def generate_xlsx_artifacts(
    input_path: str | Path,
    output_dir: str | Path,
    *,
    analyzed_sheets: list[str] | None = None,
    assume_full_analysis: bool = False,
) -> dict[str, Path]:
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)

    workbook, stats = inspect_workbook(input_path)
    sheet_inventory = build_sheet_inventory(workbook, stats)
    row_summary = build_row_count_summary(workbook, stats)
    coverage_report = build_coverage_report(
        workbook,
        stats,
        analyzed_sheets=analyzed_sheets,
        assume_full_analysis=assume_full_analysis,
    )
    validator_result = build_validator_result(sheet_inventory, row_summary, coverage_report)

    outputs = {
        "sheet_inventory": output / "sheet_inventory.json",
        "row_count_summary": output / "row_count_summary.json",
        "coverage_report": output / "coverage_report.json",
        "validator_result": output / "validator_result.json",
    }
    locked_atomic_write_json(outputs["sheet_inventory"], sheet_inventory)
    locked_atomic_write_json(outputs["row_count_summary"], row_summary)
    locked_atomic_write_json(outputs["coverage_report"], coverage_report)
    locked_atomic_write_json(outputs["validator_result"], validator_result)
    return outputs


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Generate MCPGuardian XLSX validation artifacts")
    parser.add_argument("--input", required=True, help="Path to .xlsx or .xlsm workbook")
    parser.add_argument("--output-dir", required=True, help="Run directory where artifacts will be written")
    parser.add_argument(
        "--analyzed-sheet",
        action="append",
        default=[],
        help="Sheet name declared as analytically covered. Repeatable. Use '*' for all sheets.",
    )
    parser.add_argument(
        "--assume-full-analysis",
        action="store_true",
        help="Declare all workbook sheets/rows/cells analytically covered. Use only after full analysis.",
    )
    args = parser.parse_args(argv)

    outputs = generate_xlsx_artifacts(
        args.input,
        args.output_dir,
        analyzed_sheets=args.analyzed_sheet,
        assume_full_analysis=args.assume_full_analysis,
    )
    for name, path in outputs.items():
        print(f"{name}: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
